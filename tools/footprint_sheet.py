# -*- coding: utf-8 -*-
"""`... - PCB Footprint 현황(YYMMDD).xlsx` 검토·수정.

    python footprint_sheet.py <sheet.xlsx> check   [--pcbdoc <file.PcbDoc>]
    python footprint_sheet.py <sheet.xlsx> diff    <이전.xlsx>
    python footprint_sheet.py <sheet.xlsx> formula [--apply]

Sheet1 은 부품 그룹(C Part Reference / D Value / E PCB Footprint / F Size),
Sheet2 는 지정자 단위 전개(A 지정자 / B Value / C PCB Footprint / D 검토 결과).

풋프린트의 정답은 이 엑셀이 아니라 **PcbDoc 의 실제 배치**다. `check --pcbdoc`
가 그 대조를 한다. 엑셀이 틀린 경우도(E5/E6 이 BLE 인데 915MHz 안테나로 기재),
PcbDoc 이 회로 변경을 아직 못 따라간 경우도(J1 이 구부품인 채로 배치) 둘 다
나온다. 어느 쪽인지는 Value 를 보고 사람이 판단한다.

Sheet2 C 열은 **지정자 기준**으로 조회해야 한다. Value 기준 VLOOKUP 은 값이 같고
패키지가 다른 부품에서 첫 일치행만 돌려주므로 조용히 틀린다 - `0R/1%` 한 그룹
때문에 39건이 L0402 로 잘못 나온 적이 있다. `formula` 가 내주는 수식은 콤마로
묶인 지정자 목록을 정확히 매칭하고, 멀티파트 심볼(U5A/U5B/U5C 는 Sheet1 에 U5
하나뿐)도 끝 글자를 떼고 재시도해 받아낸다.
"""
import argparse
import os
import re
import shutil
import sys
import time
import zipfile

import openpyxl
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

TXT = lambda v: '' if v is None else str(v).strip()   # noqa: E731


def designators(cell):
    return [t.strip() for t in str(cell or '').split(',') if t.strip()]


def sort_key(des):
    digits = re.sub(r'\D', '', des)
    return (re.sub(r'\d', '', des), int(digits) if digits else 0)


def load(path):
    """Sheet1 -> {지정자: (행, Value, Footprint, Size)}"""
    if not os.path.exists(path):
        raise SystemExit('파일이 없다: %s' % path)
    folder, name = os.path.dirname(path), os.path.basename(path)
    if os.path.exists(os.path.join(folder, '~$' + name)):
        print('  (참고: %s 가 Excel 에서 열려 있다. 읽기는 되지만 저장 전 값일 수 있다)' % name)
    try:
        book = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        raise SystemExit('열지 못했다: %s\n  %s\n  구형 .xls 는 .xlsx 로 저장한 뒤 쓸 것'
                         % (path, exc))
    if 'Sheet1' not in book.sheetnames:
        raise SystemExit('Sheet1 이 없다. 시트: %s' % book.sheetnames)
    ws = book['Sheet1']
    by = {}
    for r in range(2, ws.max_row + 1):
        ref = ws.cell(r, 3).value
        if not ref:
            continue
        for x in designators(ref):
            by[x] = (r, TXT(ws.cell(r, 4).value), TXT(ws.cell(r, 5).value), TXT(ws.cell(r, 6).value))
    return ws, by


def board_footprints(pcbdoc):
    from parse_pcbdoc import components
    return components(pcbdoc)


def cmd_check(path, pcbdoc):
    ws, by = load(path)
    print('%s — Sheet1 지정자 %d개' % (os.path.basename(path), len(by)))

    blank = sorted([x for x, (_, _, fp, _) in by.items() if not fp], key=sort_key)
    print('\n  풋프린트 공란: %d개 %s' % (len(blank), blank[:20] if blank else ''))

    groups = {}
    for x, (_, val, fp, _) in by.items():
        if val and fp:
            groups.setdefault(val, set()).add(fp)
    clash = {k: v for k, v in groups.items() if len(v) > 1}
    print('  Value 는 같은데 풋프린트가 다른 그룹: %d건' % len(clash))
    for val, fps in clash.items():
        print('     %-24s %s' % (val, ' / '.join(sorted(fps))))
    if clash:
        print('     -> Sheet2 는 반드시 지정자 기준 수식이어야 한다 (formula 참조)')

    try:
        s2 = openpyxl.load_workbook(path, data_only=True)['Sheet2']
        errs = [(r, TXT(s2.cell(r, 1).value), s2.cell(r, 3).value) for r in range(2, s2.max_row + 1)
                if isinstance(s2.cell(r, 3).value, str) and s2.cell(r, 3).value.startswith('#')]
        print('\n  Sheet2 수식 오류: %d개 %s'
              % (len(errs), [(a, v) for _, a, v in errs[:10]] if errs else ''))
        missing = sorted({TXT(s2.cell(r, 1).value) for r in range(2, s2.max_row + 1)
                          if TXT(s2.cell(r, 1).value)
                          and TXT(s2.cell(r, 1).value) not in by
                          and TXT(s2.cell(r, 1).value)[:-1] not in by}, key=sort_key)
        print('  Sheet1 에 없는 Sheet2 지정자: %d개 %s' % (len(missing), missing[:20]))
    except KeyError:
        pass

    if not pcbdoc:
        print('\n  (--pcbdoc 을 주면 실제 배치와 대조한다)')
        return
    board = board_footprints(pcbdoc)
    print('\n=== PcbDoc 대조 (%d개 배치) ===' % len(board))
    only_sheet = sorted(set(by) - set(board), key=sort_key)
    only_board = sorted(set(board) - set(by), key=sort_key)
    diff = [(x, by[x][2], board[x]) for x in sorted(set(by) & set(board), key=sort_key)
            if by[x][2] and by[x][2] != board[x]]
    print('  엑셀에만: %d개 %s' % (len(only_sheet), only_sheet[:20]))
    print('  PcbDoc 에만: %d개 %s' % (len(only_board), only_board[:20]))
    print('  풋프린트 불일치: %d건' % len(diff))
    for x, sheet_fp, board_fp in diff:
        same = sheet_fp.lower() == board_fp.lower()
        print('     %-8s 엑셀=%-28s PcbDoc=%-28s%s'
              % (x, sheet_fp, board_fp, '  (대소문자 차이)' if same else ''))


def cmd_diff(new_path, old_path):
    _, new = load(new_path)
    _, old = load(old_path)
    print('%s  vs  %s' % (os.path.basename(new_path), os.path.basename(old_path)))
    added = sorted(set(new) - set(old), key=sort_key)
    removed = sorted(set(old) - set(new), key=sort_key)
    print('  신규 %d개 %s' % (len(added), added))
    print('  삭제 %d개 %s' % (len(removed), removed))
    print('\n  %-8s %-26s %-26s %s' % ('지정자', 'Value', 'Footprint', 'Size'))
    n = 0
    for x in sorted(set(new) & set(old), key=sort_key):
        _, vn, fn, sn = new[x]
        _, vo, fo, so = old[x]
        if (vn, fn, sn) == (vo, fo, so):
            continue
        n += 1
        mark = lambda a, b: '%s → %s' % (a, b) if a != b else a      # noqa: E731
        print('  %-8s %-26s %-26s %s' % (x, mark(vo, vn)[:25], mark(fo, fn)[:25], mark(so, sn)))
    print('  변경 %d건' % n)


def build_formula(row, last_row):
    base = ('LOOKUP(2,1/ISNUMBER(SEARCH(","&{k}&",",","&SUBSTITUTE(Sheet1!$C$1:$C${L}," ","")&",")),'
            'Sheet1!$E$1:$E${L})')
    exact = base.format(k='A%d' % row, L=last_row)
    fallback = base.format(k='LEFT(A{0},LEN(A{0})-1)'.format(row), L=last_row)
    return '=IFERROR(%s,%s)' % (exact, fallback)


def cmd_formula(path, apply_it):
    ws, by = load(path)
    last = ws.max_row
    print('Sheet1 %d행 기준 수식 (Sheet2 C2):' % last)
    print(build_formula(2, last))
    if not apply_it:
        print('\n  --apply 를 주면 Sheet2 C 열 전체에 넣는다')
        return

    folder, name = os.path.dirname(path), os.path.basename(path)
    if os.path.exists(os.path.join(folder, '~$' + name)):
        print('\n  이 파일이 Excel 에서 열려 있다. 닫고 다시 실행할 것'); return
    risky = [x for x in zipfile.ZipFile(path).namelist()
             if any(k in x.lower() for k in ('chart', 'drawing', 'media', 'pivot', 'comments'))]
    if risky:
        print('\n  openpyxl 왕복에서 손실될 요소가 있다: %s' % risky); return
    backup = path.replace('.xlsx', '.bak_%s.xlsx' % time.strftime('%Y%m%d'))
    shutil.copy2(path, backup)

    wb = openpyxl.load_workbook(path)
    s2 = wb['Sheet2']
    count = 0
    for r in range(2, s2.max_row + 1):
        if not s2.cell(r, 1).value:
            continue
        s2.cell(r, 3).value = build_formula(r, last)
        count += 1
    wb.save(path)

    unresolved = [TXT(s2.cell(r, 1).value) for r in range(2, s2.max_row + 1)
                  if TXT(s2.cell(r, 1).value)
                  and TXT(s2.cell(r, 1).value) not in by
                  and TXT(s2.cell(r, 1).value)[:-1] not in by]
    print('\n  백업 %s' % os.path.basename(backup))
    print('  C 열 %d행 갱신' % count)
    print('  폴백 후에도 해결 안 되는 지정자: %s' % (unresolved or '없음'))
    print('  Excel 에서 열어 재계산할 것 (안 바뀌면 Ctrl+Alt+F9)')


def main():
    ap = argparse.ArgumentParser(add_help=False)
    ap.add_argument('sheet')
    ap.add_argument('command', choices=['check', 'diff', 'formula'])
    ap.add_argument('other', nargs='?')
    ap.add_argument('--pcbdoc')
    ap.add_argument('--apply', action='store_true')
    if len(sys.argv) < 3:
        print(__doc__)
        return 1
    a = ap.parse_args()

    if a.command == 'check':
        cmd_check(a.sheet, a.pcbdoc)
    elif a.command == 'diff':
        if not a.other:
            print('비교할 이전 파일을 지정할 것'); return 1
        cmd_diff(a.sheet, a.other)
    else:
        cmd_formula(a.sheet, a.apply)
    return 0


if __name__ == '__main__':
    sys.exit(main())
